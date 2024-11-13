import openpyxl
import os
import re
import requests
import time
from collections import defaultdict
from flask import Flask, request, jsonify
from werkzeug.utils import secure_filename

# Azure OpenAI API Configuration
AZURE_API_KEY = "9c1a8daa826c4abab74f38cff791f231"
AZURE_API_BASE = "https://myfirstopenaiplayground.openai.azure.com/"
DEPLOYMENT_NAME = "gpt-35-turbo-16k"
API_VERSION = "2024-06-01"

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'converted_sheets'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

# Define headers for the Azure OpenAI API request
headers = {
    "Content-Type": "application/json",
    "api-key": AZURE_API_KEY
}

# Regular expression to normalize formulas by replacing cell references with placeholders
cell_ref_pattern = re.compile(r'\b[A-Z]+[0-9]+\b')

# Function to normalize formulas by replacing cell references with a placeholder
def normalize_formula(formula):
    return cell_ref_pattern.sub("<cell_ref>", formula)

# Function to send conversion request to LLM
def get_converted_code(formula_patterns, output_language):
    prompt = f"Generate {output_language} methods to convert the following Excel formula templates into {output_language} code with parameters: {', '.join(formula_patterns)}"
    data = {
        "messages": [
            {"role": "system", "content": f"You are an AI that generates parameterized {output_language} code from Excel formula templates."},
            {"role": "user", "content": prompt}
        ],
        "max_tokens": 200,
        "temperature": 0.7
    }
    
    url = f"{AZURE_API_BASE}/openai/deployments/{DEPLOYMENT_NAME}/chat/completions?api-version={API_VERSION}"
    
    max_retries = 5
    retry_delay = 10

    for attempt in range(max_retries):
        try:
            response = requests.post(url, headers=headers, json=data)
            response.raise_for_status()
            return response.json()['choices'][0]['message']['content'].strip()
        except requests.exceptions.RequestException as e:
            if response.status_code == 429 and attempt < max_retries - 1:
                print(f"Rate limit exceeded. Retrying in {retry_delay} seconds...")
                time.sleep(retry_delay)
                retry_delay *= 2
            else:
                print(f"Error communicating with OpenAI API: {e}")
                return None

# Function to extract, group, and process formulas from a sheet
def process_sheet(sheet, output_folder, output_language):
    sheet_name = sheet.title
    output_file_path = os.path.join(output_folder, f"{sheet_name}.{output_language.lower()}")
    
    formula_groups = defaultdict(list)  # Dictionary to group formulas by their normalized structure

    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Detect formula cells
                original_formula = cell.value
                normalized_formula = normalize_formula(original_formula)
                formula_groups[normalized_formula].append(original_formula)
    
    # Process each unique formula group
    for formula_pattern, formulas in formula_groups.items():
        if len(formulas) > 0:
            converted_code = get_converted_code([formula_pattern], output_language)
            if converted_code:
                with open(output_file_path, 'a') as file:
                    file.write(f"# Parameterized method for formulas like {formula_pattern}\n")
                    file.write(converted_code + "\n\n")
            time.sleep(10)

# Main function to process the workbook
def convert_excel_formulas_to_code(excel_file, output_language):
    workbook = openpyxl.load_workbook(excel_file, data_only=False)
    output_folder = app.config['OUTPUT_FOLDER']
    for sheet_name in workbook.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        process_sheet(sheet, output_folder, output_language)
    print("Conversion completed.")

# Flask route for the endpoint
@app.route('/excelconvert', methods=['POST'])
def excel_convert():
    # Check if a file and output language were provided
    if 'file' not in request.files or 'output_language' not in request.form:
        return jsonify({"error": "No file or output language provided."}), 400

    file = request.files['file']
    output_language = request.form['output_language']

    # Validate file type
    if file.filename == '' or not file.filename.endswith('.xlsx'):
        return jsonify({"error": "Invalid file type. Please upload an .xlsx file."}), 400

    filename = secure_filename(file.filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)

    # Convert the Excel file
    try:
        convert_excel_formulas_to_code(file_path, output_language)
        return jsonify({"message": "Conversion completed successfully.", "output_folder": app.config['OUTPUT_FOLDER']})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)
